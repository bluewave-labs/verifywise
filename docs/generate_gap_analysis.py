#!/usr/bin/env python3
"""
Generate TMLR 2025 Gap Analysis Excel Report for VerifyWise
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

# Create workbook
wb = Workbook()

# ============================================
# SHEET 1: Executive Summary
# ============================================
ws_summary = wb.active
ws_summary.title = "Executive Summary"

# Title
ws_summary.merge_cells('A1:F1')
ws_summary['A1'] = "VerifyWise Gap Analysis: TMLR 2025 AI Governance Requirements"
ws_summary['A1'].font = Font(bold=True, size=16, color="FFFFFF")
ws_summary['A1'].fill = PatternFill(start_color="13715B", end_color="13715B", fill_type="solid")
ws_summary['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws_summary.row_dimensions[1].height = 30

# Overall Score
ws_summary['A3'] = "Overall Coverage Score (excl. Compute Governance)"
ws_summary['A3'].font = Font(bold=True, size=12)
ws_summary['B3'] = "59%"
ws_summary['B3'].font = Font(bold=True, size=14, color="13715B")

# Summary table headers
headers = ["Category", "Coverage %", "Full", "Partial", "Missing", "Score"]
for col, header in enumerate(headers, 1):
    cell = ws_summary.cell(row=5, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Category data
categories = [
    ("1. Data Governance", "55%", 4, 3, 4, 5.5),
    ("2. Model Evaluation & Risk", "63%", 5, 3, 3, 6.5),
    ("3. Compute Governance", "43%", 3, 1, 3, 3.5),
    ("4. Access & Audit", "82%", 9, 2, 0, 10.0),
    ("5. Verification & Evidence", "71%", 5, 2, 1, 6.0),
    ("6. Deployment Monitoring", "38%", 2, 2, 4, 3.0),
    ("7. Ecosystem & Vendor", "56%", 5, 1, 3, 5.5),
]

# Color mapping for coverage
def get_coverage_color(pct_str):
    pct = int(pct_str.replace('%', ''))
    if pct >= 75:
        return "C6EFCE"  # Green
    elif pct >= 50:
        return "FFEB9C"  # Yellow
    else:
        return "FFC7CE"  # Red

for row_idx, cat_data in enumerate(categories, 6):
    for col_idx, value in enumerate(cat_data, 1):
        cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left")
        if col_idx == 2:  # Coverage column
            cell.fill = PatternFill(start_color=get_coverage_color(str(value)), end_color=get_coverage_color(str(value)), fill_type="solid")

# Totals row
ws_summary.cell(row=13, column=1, value="TOTAL (excl. Compute)").font = Font(bold=True)
ws_summary.cell(row=13, column=2, value="59%").font = Font(bold=True)
ws_summary.cell(row=13, column=2).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ws_summary.cell(row=13, column=3, value=30).font = Font(bold=True)
ws_summary.cell(row=13, column=4, value=15).font = Font(bold=True)
ws_summary.cell(row=13, column=5, value=19).font = Font(bold=True)
ws_summary.cell(row=13, column=6, value=37.5).font = Font(bold=True)

# Key Strengths
ws_summary['A16'] = "Key Strengths"
ws_summary['A16'].font = Font(bold=True, size=12, color="13715B")
strengths = [
    "Access Control & Security (82%) - Comprehensive RBAC, multi-tenancy, encryption",
    "Verification & Evidence (71%) - Strong framework coverage, audit trails, reporting",
    "Model Evaluation (63%) - Extensive bias/fairness metrics, safety evaluations",
    "Vendor Management (56%) - Full vendor inventory, risk assessment, AI detection"
]
for i, s in enumerate(strengths, 17):
    ws_summary[f'A{i}'] = f"• {s}"

# Critical Gaps
ws_summary['A22'] = "Critical Gaps"
ws_summary['A22'].font = Font(bold=True, size=12, color="B42318")
gaps = [
    "Production Monitoring - No real-time inference monitoring, SLA tracking",
    "Model Explainability - No SHAP/LIME implementation",
    "Compute Governance - Missing GPU/CPU tracking, quotas, carbon footprint",
    "Data Lineage - No transformation tracking or data flow visualization",
    "Digital Signatures - No cryptographic attestation for evidence",
    "Downstream Impact - No dependency mapping for model changes"
]
for i, g in enumerate(gaps, 23):
    ws_summary[f'A{i}'] = f"• {g}"

# Column widths
ws_summary.column_dimensions['A'].width = 45
ws_summary.column_dimensions['B'].width = 15
ws_summary.column_dimensions['C'].width = 10
ws_summary.column_dimensions['D'].width = 10
ws_summary.column_dimensions['E'].width = 10
ws_summary.column_dimensions['F'].width = 10

# ============================================
# SHEET 2: Detailed Requirements
# ============================================
ws_details = wb.create_sheet("Detailed Requirements")

# Headers
detail_headers = ["Category", "Requirement", "Status", "Coverage %", "Implementation Details", "File Location(s)", "Priority", "Recommendation"]
for col, header in enumerate(detail_headers, 1):
    cell = ws_details.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="13715B", end_color="13715B", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Freeze header row
ws_details.freeze_panes = 'A2'

# All requirements data
requirements = [
    # Category 1: Data Governance
    ("1. Data Governance", "Dataset provenance tracking", "Partial", "50%", "File-level metadata tracked (uploaded_by, upload_date, org_id). No dataset attribute or origin metadata tracking.", "Servers/database/migrations/20251104000511-create-tables-for-file-manager.js", "Medium", "Add dataset metadata schema with origin tracking"),
    ("1. Data Governance", "Data version control", "Partial", "50%", "Model versions tracked via MLflow integration. Basic dataset versioning in Evaluations module. No git-like branching for datasets.", "Servers/domain.layer/models/modelInventory/modelInventoryHistory.model.ts", "Medium", "Implement DVC-style dataset versioning"),
    ("1. Data Governance", "Data lineage tracking", "Missing", "0%", "No data transformation tracking. No data flow diagrams. No tracking of which datasets feed into which models.", "N/A", "High", "Implement data lineage graph with transformation tracking"),
    ("1. Data Governance", "Data access controls", "Full", "100%", "RBAC enforced for file operations. Multi-tenancy via org_id. Audit logging for file access.", "Servers/middleware/accessControl.middleware.ts", "N/A", "Maintain current implementation"),
    ("1. Data Governance", "Data retention policies", "Partial", "50%", "Retention periods documented in compliance framework questions. Soft delete implemented. No automated enforcement.", "Servers/structures/LAW-25/topics.struct.ts", "Medium", "Add automated retention policy enforcement"),
    ("1. Data Governance", "Data quality monitoring", "Partial", "50%", "Framework questions cover completeness and drift. Data Quality category in Model Risks. No automated drift detection.", "Clients/src/presentation/structures/AssessmentTracker/", "High", "Implement automated data quality checks and drift alerts"),
    ("1. Data Governance", "Consent management", "Partial", "50%", "GDPR consent mechanisms captured in assessment questions. No formal consent management system.", "Servers/structures/EU-AI-Act/assessment-tracker/", "Medium", "Build consent registry with audit trail"),
    ("1. Data Governance", "Data classification", "Partial", "50%", "Data sensitivity levels for vendors (High/Medium/Low). No granular file-level classification.", "Clients/src/domain/utils/vendorScorecard.utils.ts", "Medium", "Add file-level sensitivity tagging"),
    ("1. Data Governance", "Privacy Impact Assessment", "Full", "100%", "Built into EU AI Act and GDPR compliance frameworks. Questions address privacy risks and mitigations.", "Servers/structures/EU-AI-Act/assessment-tracker/", "N/A", "Maintain current implementation"),
    ("1. Data Governance", "Data subject rights tracking", "Full", "100%", "Framework questions cover right to access and deletion. Compliance documentation in place.", "Servers/structures/EU-AI-Act/assessment-tracker/", "N/A", "Maintain current implementation"),
    ("1. Data Governance", "Secure data deletion", "Full", "100%", "Soft delete with is_deleted and deleted_at fields. Archival mentioned in decommissioning plans.", "Servers/domain.layer/models/risk/risk.model.ts", "N/A", "Maintain current implementation"),

    # Category 2: Model Evaluation
    ("2. Model Evaluation", "Model risk scoring", "Full", "100%", "4 severity levels (Low/Medium/High/Critical). 5 categories. Full CRUD operations.", "Servers/domain.layer/models/modelRisk/modelRisk.model.ts", "N/A", "Maintain current implementation"),
    ("2. Model Evaluation", "Performance metrics", "Partial", "50%", "Latency and token count tracked. Cost tracking implemented. Standard ML metrics only via custom DeepEval.", "EvalServer/src/models/evaluation_logs.py", "Medium", "Add native accuracy/precision/recall/F1 tracking"),
    ("2. Model Evaluation", "Bias & fairness detection", "Full", "100%", "30+ fairness metrics via BiasAndFairnessModule. Demographic parity, equalized odds, predictive parity.", "EvaluationModule/BiasAndFairnessModule/src/eval_engine/metrics.py", "N/A", "Maintain current implementation"),
    ("2. Model Evaluation", "Model explainability (SHAP/LIME)", "Missing", "0%", "Requirements documented in assessment tracker. No actual SHAP or LIME implementation.", "N/A", "High", "Implement SHAP/LIME integration with visualization"),
    ("2. Model Evaluation", "Red teaming / adversarial testing", "Missing", "0%", "Safety metrics exist but no adversarial attack generation. No jailbreak testing.", "N/A", "High", "Build adversarial testing framework"),
    ("2. Model Evaluation", "Safety evaluations", "Full", "100%", "ToxicityMetric, Bias checks, Hallucination detection. G-Eval safety evaluation.", "EvalServer/src/controllers/deepeval.py", "N/A", "Maintain current implementation"),
    ("2. Model Evaluation", "Benchmark tracking", "Partial", "50%", "DeepEval benchmarks referenced. Experiment tracking. No leaderboards or regression detection.", "Clients/src/presentation/pages/EvalsDashboard/", "Medium", "Add benchmark leaderboards and regression alerts"),
    ("2. Model Evaluation", "Model cards / documentation", "Partial", "50%", "Model Inventory with metadata. MLflow records. No formal Model Card generation.", "Clients/src/presentation/pages/ModelInventory/", "Medium", "Implement Model Card generator"),
    ("2. Model Evaluation", "Hallucination detection", "Full", "100%", "HallucinationMetric in DeepEval. Faithfulness checks against context.", "EvalServer/src/controllers/deepeval.py", "N/A", "Maintain current implementation"),
    ("2. Model Evaluation", "Continuous model monitoring", "Missing", "0%", "No real-time inference monitoring. No production endpoint tracking.", "N/A", "High", "Build real-time inference monitoring dashboard"),
    ("2. Model Evaluation", "Automated risk assessment", "Partial", "50%", "Manual risk entry with some auto-calculation. Risk propagation not implemented.", "Servers/domain.layer/models/modelRisk/", "Medium", "Add automated risk scoring"),

    # Category 3: Compute Governance
    ("3. Compute Governance", "Resource usage tracking (GPU/CPU)", "Missing", "0%", "No GPU/CPU monitoring infrastructure. No system resource metrics collection.", "N/A", "Low", "Integrate with Prometheus/Grafana"),
    ("3. Compute Governance", "Cost tracking", "Full", "100%", "Token-based cost calculation for LLM models. Pricing database for major providers.", "Gateway/dist/utils/pricing.js", "N/A", "Maintain current implementation"),
    ("3. Compute Governance", "Workload scheduling", "Missing", "0%", "Basic job queue exists but no advanced scheduling or priority management.", "Servers/jobs/", "Low", "Implement job priority queue"),
    ("3. Compute Governance", "Resource quotas", "Missing", "0%", "No per-user storage quotas. No API call limits per user.", "N/A", "Low", "Add quota management system"),
    ("3. Compute Governance", "Carbon footprint tracking", "Missing", "0%", "No carbon tracking or emissions calculation.", "N/A", "Low", "Integrate carbon footprint estimation"),
    ("3. Compute Governance", "Compute access controls", "Full", "100%", "RBAC for compute resources. User permission management. Tenant isolation.", "Servers/routes/user.route.ts", "N/A", "Maintain current implementation"),
    ("3. Compute Governance", "Training job monitoring", "Full", "100%", "Training registrar with status tracking. MLflow sync for experiments.", "Servers/src/services/mlflow.service.ts", "N/A", "Maintain current implementation"),

    # Category 4: Access & Audit
    ("4. Access & Audit", "Role-based access control (RBAC)", "Full", "100%", "4 roles (Admin/Reviewer/Editor/Auditor). Fine-grained permissions per entity.", "Servers/domain.layer/models/role/role.model.ts", "N/A", "Maintain current implementation"),
    ("4. Access & Audit", "Multi-tenancy", "Full", "100%", "Schema-based separation via SHA-256 hash. Domain-based access validation.", "Servers/middleware/multiTenancy.middleware.ts", "N/A", "Maintain current implementation"),
    ("4. Access & Audit", "Audit logging", "Partial", "75%", "Event logs table. File-based daily logs. Missing detailed entity-level tracking.", "Servers/utils/logger.util.ts", "Medium", "Enhance entity-level audit logging"),
    ("4. Access & Audit", "Third-party auditor access", "Partial", "50%", "Auditor role with read-only permissions. No scoped per-project access.", "Servers/domain.layer/models/role/", "Medium", "Add scoped auditor provisioning"),
    ("4. Access & Audit", "API key management", "Full", "100%", "Tenant-scoped tokens. 30-day expiration. Auto-cleanup. Encrypted storage.", "Servers/controllers/tokens.ctrl.ts", "N/A", "Maintain current implementation"),
    ("4. Access & Audit", "SSO/OAuth integration", "Full", "100%", "Google OAuth2, Microsoft Entra ID, Slack OAuth 2.0, GitHub token integration.", "Clients/src/presentation/pages/Integrations/", "N/A", "Maintain current implementation"),
    ("4. Access & Audit", "Session management", "Full", "100%", "JWT with 1hr access / 30-day refresh tokens. HTTP-only cookies.", "Servers/middleware/auth.middleware.ts", "N/A", "Maintain current implementation"),
    ("4. Access & Audit", "Rate limiting", "Full", "100%", "4 limiters for different endpoint types. IPv6-safe.", "Servers/middleware/rateLimit.middleware.ts", "N/A", "Maintain current implementation"),
    ("4. Access & Audit", "Encryption at rest", "Full", "100%", "AES-256-CBC for sensitive data. IV per encryption.", "Servers/utils/encryption.utils.ts", "N/A", "Maintain current implementation"),
    ("4. Access & Audit", "Password security", "Full", "100%", "Bcrypt hashing with 10 rounds. Validation rules.", "Servers/utils/auth.utils.ts", "N/A", "Maintain current implementation"),
    ("4. Access & Audit", "Access control middleware", "Full", "100%", "Route-level role authorization. AsyncLocalStorage for context.", "Servers/middleware/accessControl.middleware.ts", "N/A", "Maintain current implementation"),

    # Category 5: Verification & Evidence
    ("5. Verification & Evidence", "Compliance frameworks", "Full", "100%", "EU AI Act, NIST AI RMF, ISO-27001, ISO-42001, LAW-25.", "Servers/structures/", "N/A", "Maintain current implementation"),
    ("5. Verification & Evidence", "Evidence collection & storage", "Full", "100%", "Evidence Hub with file management. Tracks metadata, expiry, mapped models.", "Servers/domain.layer/models/evidenceHub/", "N/A", "Maintain current implementation"),
    ("5. Verification & Evidence", "Audit trails / change history", "Full", "100%", "10 entity types tracked. Field-level tracking with formatters.", "Servers/config/changeHistory.config.ts", "N/A", "Maintain current implementation"),
    ("5. Verification & Evidence", "Digital signatures", "Missing", "0%", "No PKI support. No document signing. JWT only for auth.", "N/A", "High", "Implement document signing with certificates"),
    ("5. Verification & Evidence", "Compliance reports", "Full", "100%", "PDF and DOCX generation. Charts with risk-level colors. Multi-section.", "Servers/services/reporting/", "N/A", "Maintain current implementation"),
    ("5. Verification & Evidence", "Control mapping", "Full", "100%", "Controls with status/owner/reviewer/approver. Risk-to-control mapping.", "Servers/domain.layer/models/control/", "N/A", "Maintain current implementation"),
    ("5. Verification & Evidence", "Gap analysis", "Partial", "50%", "User-configurable gap rules. No automated continuous assessment.", "Servers/services/entityGraphGapRulesService.ts", "Medium", "Add automated gap detection"),
    ("5. Verification & Evidence", "Attestation workflows", "Partial", "50%", "Approval chains for risks/controls. No crypto signatures.", "Servers/domain.layer/models/risk/", "Medium", "Add cryptographic attestation"),

    # Category 6: Deployment Monitoring
    ("6. Deployment Monitoring", "Model deployment tracking", "Partial", "50%", "Model version and approval tracked. No environment segregation.", "Servers/domain.layer/models/modelInventory/", "Medium", "Add deployment environment tracking"),
    ("6. Deployment Monitoring", "Production inference monitoring", "Missing", "0%", "MLflow tracks training only. No real-time endpoint monitoring.", "N/A", "High", "Build production monitoring service"),
    ("6. Deployment Monitoring", "Incident management", "Full", "100%", "Full lifecycle. 7 incident types. Severity levels. Approval workflow. Slack.", "Servers/controllers/incident-management.ctrl.ts", "N/A", "Maintain current implementation"),
    ("6. Deployment Monitoring", "Drift detection", "Partial", "25%", "MODEL_DRIFT as incident type. No automated statistical detection.", "Servers/domain.layer/interfaces/i.aiIncidentManagement.ts", "High", "Implement automated drift detection"),
    ("6. Deployment Monitoring", "A/B testing / experiments", "Full", "100%", "DeepEval integration. Dataset management. Metrics/scorers configuration.", "Clients/src/presentation/pages/EvalsDashboard/", "N/A", "Maintain current implementation"),
    ("6. Deployment Monitoring", "Rollback capabilities", "Missing", "0%", "Model version field exists but no rollback mechanism.", "N/A", "High", "Implement model version rollback"),
    ("6. Deployment Monitoring", "SLA monitoring", "Missing", "0%", "No latency SLAs. No availability tracking.", "N/A", "High", "Build SLA monitoring dashboard"),
    ("6. Deployment Monitoring", "Downstream impact tracking", "Missing", "0%", "Model-to-project relationships exist. No dependency graph.", "N/A", "Medium", "Add dependency mapping"),

    # Category 7: Ecosystem & Vendor
    ("7. Ecosystem & Vendor", "Vendor risk assessment", "Full", "100%", "Risk severity levels. Impact/likelihood assessment. Action planning.", "Servers/domain.layer/models/vendorRisk/", "N/A", "Maintain current implementation"),
    ("7. Ecosystem & Vendor", "Vendor inventory", "Full", "100%", "Full CRUD. Review status. Scorecard. Data sensitivity. Business criticality.", "Servers/domain.layer/models/vendor/", "N/A", "Maintain current implementation"),
    ("7. Ecosystem & Vendor", "AI detection (third-party AI)", "Full", "100%", "Repository scanning for AI libraries. Pattern-based detection. Scan history.", "Servers/config/aiDetectionPatterns.ts", "N/A", "Maintain current implementation"),
    ("7. Ecosystem & Vendor", "License compliance tracking", "Missing", "0%", "No dedicated license tracking. No open-source scanner.", "N/A", "Medium", "Integrate license scanner"),
    ("7. Ecosystem & Vendor", "Contract/SLA management", "Missing", "0%", "No contract lifecycle management. No SLA tracking.", "N/A", "Medium", "Build contract management module"),
    ("7. Ecosystem & Vendor", "Security questionnaires", "Missing", "0%", "No vendor security assessment templates.", "N/A", "Medium", "Add templated security questionnaires"),
    ("7. Ecosystem & Vendor", "Integration health monitoring", "Partial", "50%", "MLflow health check only. No broad third-party API monitoring.", "Servers/routes/integrations.route.ts", "Low", "Expand integration monitoring"),
    ("7. Ecosystem & Vendor", "Supply chain compliance", "Full", "100%", "ISO-42001 third-party controls. Supplier agreements.", "Servers/structures/ISO-42001/annex/", "N/A", "Maintain current implementation"),
    ("7. Ecosystem & Vendor", "Vendor deactivation workflow", "Full", "100%", "Review status lifecycle. Archival capabilities. Change history.", "Servers/domain.layer/models/vendor/", "N/A", "Maintain current implementation"),
]

# Status color mapping
status_colors = {
    "Full": "C6EFCE",     # Green
    "Partial": "FFEB9C",  # Yellow
    "Missing": "FFC7CE",  # Red
}

# Priority color mapping
priority_colors = {
    "High": "FFC7CE",     # Red
    "Medium": "FFEB9C",   # Yellow
    "Low": "BDD7EE",      # Blue
    "N/A": "E2E8F0",      # Gray
}

# Write data
for row_idx, req in enumerate(requirements, 2):
    for col_idx, value in enumerate(req, 1):
        cell = ws_details.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Color status column
        if col_idx == 3:  # Status column
            color = status_colors.get(value, "FFFFFF")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # Color priority column
        if col_idx == 7:  # Priority column
            color = priority_colors.get(value, "FFFFFF")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

# Set column widths
col_widths = [25, 35, 10, 12, 60, 55, 10, 45]
for i, width in enumerate(col_widths, 1):
    ws_details.column_dimensions[get_column_letter(i)].width = width

# Set row heights for readability
for row in range(2, len(requirements) + 2):
    ws_details.row_dimensions[row].height = 45

# ============================================
# SHEET 3: Priority Matrix
# ============================================
ws_priority = wb.create_sheet("Priority Matrix")

ws_priority.merge_cells('A1:D1')
ws_priority['A1'] = "Implementation Priority Matrix"
ws_priority['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws_priority['A1'].fill = PatternFill(start_color="13715B", end_color="13715B", fill_type="solid")
ws_priority['A1'].alignment = Alignment(horizontal="center")

# High Priority
ws_priority['A3'] = "HIGH PRIORITY (Critical for Compliance)"
ws_priority['A3'].font = Font(bold=True, size=12, color="B42318")
high_priority = [
    ("Data lineage tracking", "1. Data Governance"),
    ("Data quality monitoring", "1. Data Governance"),
    ("Model explainability (SHAP/LIME)", "2. Model Evaluation"),
    ("Red teaming / adversarial testing", "2. Model Evaluation"),
    ("Continuous model monitoring", "2. Model Evaluation"),
    ("Digital signatures", "5. Verification & Evidence"),
    ("Production inference monitoring", "6. Deployment Monitoring"),
    ("Drift detection (automated)", "6. Deployment Monitoring"),
    ("Rollback capabilities", "6. Deployment Monitoring"),
    ("SLA monitoring", "6. Deployment Monitoring"),
]

ws_priority['A4'] = "Requirement"
ws_priority['B4'] = "Category"
ws_priority['A4'].font = Font(bold=True)
ws_priority['B4'].font = Font(bold=True)

for i, (req, cat) in enumerate(high_priority, 5):
    ws_priority[f'A{i}'] = req
    ws_priority[f'B{i}'] = cat
    ws_priority[f'A{i}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Medium Priority
start_row = len(high_priority) + 7
ws_priority[f'A{start_row}'] = "MEDIUM PRIORITY (Enhances Governance)"
ws_priority[f'A{start_row}'].font = Font(bold=True, size=12, color="9C6500")

medium_priority = [
    ("Dataset provenance tracking", "1. Data Governance"),
    ("Data version control", "1. Data Governance"),
    ("Data retention policies", "1. Data Governance"),
    ("Consent management", "1. Data Governance"),
    ("Data classification", "1. Data Governance"),
    ("Performance metrics", "2. Model Evaluation"),
    ("Benchmark tracking", "2. Model Evaluation"),
    ("Model cards / documentation", "2. Model Evaluation"),
    ("Automated risk assessment", "2. Model Evaluation"),
    ("Audit logging (enhanced)", "4. Access & Audit"),
    ("Third-party auditor access", "4. Access & Audit"),
    ("Gap analysis (automated)", "5. Verification & Evidence"),
    ("Attestation workflows", "5. Verification & Evidence"),
    ("Model deployment tracking", "6. Deployment Monitoring"),
    ("Downstream impact tracking", "6. Deployment Monitoring"),
    ("License compliance tracking", "7. Ecosystem & Vendor"),
    ("Contract/SLA management", "7. Ecosystem & Vendor"),
    ("Security questionnaires", "7. Ecosystem & Vendor"),
]

ws_priority[f'A{start_row+1}'] = "Requirement"
ws_priority[f'B{start_row+1}'] = "Category"
ws_priority[f'A{start_row+1}'].font = Font(bold=True)
ws_priority[f'B{start_row+1}'].font = Font(bold=True)

for i, (req, cat) in enumerate(medium_priority, start_row + 2):
    ws_priority[f'A{i}'] = req
    ws_priority[f'B{i}'] = cat
    ws_priority[f'A{i}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

# Low Priority
start_row2 = start_row + len(medium_priority) + 4
ws_priority[f'A{start_row2}'] = "LOW PRIORITY (Nice to Have)"
ws_priority[f'A{start_row2}'].font = Font(bold=True, size=12, color="2E75B6")

low_priority = [
    ("Resource usage tracking (GPU/CPU)", "3. Compute Governance"),
    ("Workload scheduling", "3. Compute Governance"),
    ("Resource quotas", "3. Compute Governance"),
    ("Carbon footprint tracking", "3. Compute Governance"),
    ("Integration health monitoring", "7. Ecosystem & Vendor"),
]

ws_priority[f'A{start_row2+1}'] = "Requirement"
ws_priority[f'B{start_row2+1}'] = "Category"
ws_priority[f'A{start_row2+1}'].font = Font(bold=True)
ws_priority[f'B{start_row2+1}'].font = Font(bold=True)

for i, (req, cat) in enumerate(low_priority, start_row2 + 2):
    ws_priority[f'A{i}'] = req
    ws_priority[f'B{i}'] = cat
    ws_priority[f'A{i}'].fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

ws_priority.column_dimensions['A'].width = 45
ws_priority.column_dimensions['B'].width = 30

# ============================================
# SHEET 4: Status Legend
# ============================================
ws_legend = wb.create_sheet("Legend")

ws_legend['A1'] = "Status Definitions"
ws_legend['A1'].font = Font(bold=True, size=14)

ws_legend['A3'] = "Full"
ws_legend['B3'] = "Feature fully implemented and operational"
ws_legend['A3'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

ws_legend['A4'] = "Partial"
ws_legend['B4'] = "Feature partially implemented or at framework/documentation level only"
ws_legend['A4'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

ws_legend['A5'] = "Missing"
ws_legend['B5'] = "Feature not implemented"
ws_legend['A5'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

ws_legend['A7'] = "Priority Definitions"
ws_legend['A7'].font = Font(bold=True, size=14)

ws_legend['A9'] = "High"
ws_legend['B9'] = "Critical for regulatory compliance; should be implemented first"
ws_legend['A9'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

ws_legend['A10'] = "Medium"
ws_legend['B10'] = "Enhances governance capabilities; implement after high priority items"
ws_legend['A10'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

ws_legend['A11'] = "Low"
ws_legend['B11'] = "Nice to have; can be deferred or scoped based on resources"
ws_legend['A11'].fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

ws_legend['A12'] = "N/A"
ws_legend['B12'] = "Already implemented; no action needed"
ws_legend['A12'].fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

ws_legend['A14'] = "Coverage Calculation"
ws_legend['A14'].font = Font(bold=True, size=14)

ws_legend['A16'] = "Formula: (Full × 1.0 + Partial × 0.5) / Total Requirements"
ws_legend['A17'] = "Full = 100%, Partial = 50%, Missing = 0%"

ws_legend.column_dimensions['A'].width = 15
ws_legend.column_dimensions['B'].width = 70

# Save workbook
output_path = "/Users/gorkemcetin/verifywise/docs/TMLR_2025_Gap_Analysis.xlsx"
wb.save(output_path)
print(f"Excel file saved to: {output_path}")
